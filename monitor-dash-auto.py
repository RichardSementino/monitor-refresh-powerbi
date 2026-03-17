import io
import os
import json
import openpyxl
from datetime import datetime
import locale
import requests
from azure.storage.blob import BlobServiceClient
from dotenv import load_dotenv

load_dotenv()


# ==============================
# CONFIGURAÇÃO DE LOCALE
# ==============================

try:
    locale.setlocale(locale.LC_TIME, "portuguese_brazil")
except:
    locale.setlocale(locale.LC_TIME, "pt_BR.utf8")

class MonitorDadosEPTV:
    def __init__(self, conn_str, container, arquivo):
        self.service_client = BlobServiceClient.from_connection_string(conn_str)
        self.blob_client = self.service_client.get_blob_client(container=container, blob=arquivo)

        self.wb = None
        self.ws_monitor = None
        self.ws_erros = None

    # ==============================
    # DOWNLOAD ÚNICO DO EXCEL
    # ==============================
    def carregar_workbook(self):
        print("📥 Baixando arquivo do Blob...")
        stream = self.blob_client.download_blob().readall()
        file_buffer = io.BytesIO(stream)
        self.wb = openpyxl.load_workbook(file_buffer)
        self.ws_monitor = self.wb["Monitoramento Diário"]
        self.ws_erros = self.wb["Erros Dashboards"]

    # ==============================
    # UPLOAD DO EXCEL
    # ==============================
    def salvar_workbook(self):
        print("📤 Enviando arquivo atualizado ao Blob...")
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        self.blob_client.upload_blob(output, overwrite=True)
        print("✅ Upload concluído com sucesso.")

    # ==============================
    # TOKEN POWER BI
    # ==============================
    def gerar_token_pbi(self, tenant_id, client_id, client_secret):
        print("🔑 Gerando novo Token Power BI...")

        url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        body = {
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://analysis.windows.net/powerbi/api/.default"
        }

        try:
            response = requests.post(url, data=body)
            response.raise_for_status()
            return response.json().get("access_token")
        except Exception as e:
            print(f"❌ Erro ao gerar token: {e}")
            return None

    # ==============================
    # STATUS + DETALHE REFRESH
    # ==============================
    def obter_status_e_detalhe(self, workspace_id, dataset_id, token):
        url = f"https://api.powerbi.com/v1.0/myorg/groups/{workspace_id}/datasets/{dataset_id}/refreshes?$top=1"

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }

        try:
            response = requests.get(url, headers=headers)
            if response.status_code != 200:
                return "Erro", f"HTTP {response.status_code}: {response.text}", None, None

            dados = response.json()
            if not dados.get("value"):
                return "Erro", "Sem histórico de refresh.", None, None

            ultimo = dados["value"][0]
            status = ultimo.get("status")

            start = ultimo.get("startTime")
            end = ultimo.get("endTime")
            
            refresh_time = end if end else start
            
            if refresh_time:
                refresh_dt = datetime.fromisoformat(refresh_time.replace("Z", ""))
                data_refresh = refresh_dt.strftime("%d/%m/%Y")
                hora_refresh = refresh_dt.strftime("%H:%M")
            else:
                data_refresh = None
                hora_refresh = None
                
            if status == "Completed":
                return "OK", None, data_refresh, hora_refresh

            # Extrair detalhe
            detalhe = ultimo.get("serviceExceptionJson")
            start = ultimo.get("startTime")
            end = ultimo.get("endTime")

            msg = []
            msg.append(f"Dataset: {dataset_id}")
            msg.append(f"Status: {status}")
            if start: msg.append(f"Início: {start}")
            if end: msg.append(f"Fim: {end}")

            if detalhe:
                try:
                    detalhe = json.loads(detalhe)
                except:
                    pass
                msg.append("Detalhe:")
                msg.append(str(detalhe))

            return "Erro", "\n".join(msg), data_refresh, hora_refresh

        except Exception as e:
            return "Erro", f"Falha comunicação API: {e}", None, None

    # ==============================
    # ATUALIZA MONITORAMENTO
    # ==============================
    def atualizar_monitoramento(self, painel, status):
        hoje_dt = datetime.now().date()
        hoje_pt = hoje_dt.strftime("%d/%b/%y").lower()

        row_idx, col_idx = None, None

        # Buscar linha da data
        for r in range(5, self.ws_monitor.max_row + 1):
            val = self.ws_monitor.cell(row=r, column=1).value
            if val is None:
                continue

            if isinstance(val, datetime):
                if val.date() == hoje_dt:
                    row_idx = r
                    break
            else:
                if str(val).strip().lower().replace("-", "/") == hoje_pt:
                    row_idx = r
                    break

        # Buscar coluna do painel
        for c in range(2, self.ws_monitor.max_column + 1):
            header = self.ws_monitor.cell(row=4, column=c).value
            if header and str(header).strip() == painel:
                col_idx = c
                break

        if row_idx and col_idx:
            self.ws_monitor.cell(row=row_idx, column=col_idx).value = status
            print(f"✔ Status '{status}' atualizado para {painel}")
        else:
            print(f"⚠ Não encontrado: {painel} em {hoje_pt}")

    # ==============================
    # REGISTRAR ERRO
    # ==============================
    def registrar_erro_dashboard(self, painel, mensagem):
        if not mensagem:
            return

        hoje_dt = datetime.now().date()
        hoje_pt = hoje_dt.strftime("%d/%b/%y").lower()
        weekday_pt = hoje_dt.strftime("%A")

        # Próxima linha vazia
        next_row = self.ws_erros.max_row + 1

        self.ws_erros.cell(row=next_row, column=1).value = hoje_pt
        self.ws_erros.cell(row=next_row, column=2).value = weekday_pt
        self.ws_erros.cell(row=next_row, column=3).value = painel
        self.ws_erros.cell(row=next_row, column=4).value = mensagem

        print(f"🧾 Erro registrado para {painel}")

    def atualizar_catalogo_reports(self, painel, data_refresh, hora_refresh):
        ws_cat = self.wb["Catálogo_Reports"]
        for r in range(4, ws_cat.max_row + 1):
            nome_rel = ws_cat.cell(row=r, column=1).value
            if nome_rel and str(nome_rel).strip() == painel:
                
                ws_cat.cell(row=r, column=5).value = data_refresh    
                ws_cat.cell(row=r, column=6).value = hora_refresh
                
                print(f"📚 Catálogo atualizado → {painel}")
                break
# ==============================
# MAIN
# ==============================
if __name__ == "__main__":

    # 🔐 Variáveis de ambiente (SEGURANÇA)
    MINHA_CONEXAO = os.getenv("AZURE_BLOB_CONNECTION_STRING")
    T_ID = os.getenv("PBI_TENANT_ID")
    C_ID = os.getenv("PBI_CLIENT_ID")
    C_SECRET = os.getenv("PBI_CLIENT_SECRET")

    if not all([MINHA_CONEXAO, T_ID, C_ID, C_SECRET]):
        print("❌ Variáveis de ambiente não configuradas corretamente.")
        exit()

    estagio_robo = MonitorDadosEPTV(
        MINHA_CONEXAO,
        "monitoramento",
        "Catálogos de Gestão_Dados_copia.xlsx"
    )

    TOKEN = estagio_robo.gerar_token_pbi(T_ID, C_ID, C_SECRET)

    if not TOKEN:
        print("❌ Falha ao gerar token.")
        exit()

    # Lista reduzida aqui só exemplo
    meus_relatorios = [
            {"nome_excel": "Painel Compras", "w_id": "3ed8daac-5cc7-47f5-b4f3-adfe6deaacc7", "d_id": "071aa851-69b3-44d7-834d-e4ea20fdac71"},
            {"nome_excel": "Painel Contratos", "w_id": "3ed8daac-5cc7-47f5-b4f3-adfe6deaacc7", "d_id": "11233c49-9845-466d-a562-5c3f35fbb585"},
            {"nome_excel": "Painel Painéis Rodoviários", "w_id": "3ed8daac-5cc7-47f5-b4f3-adfe6deaacc7", "d_id": "34bcbb02-0765-439c-ad3e-a2e5c0cfb15d"},
            {"nome_excel": "Performance Campanhas", "w_id": "e12eb10e-3ad3-4583-baa8-c1b72159e21c", "d_id": "81b9d7c4-fcc4-4338-9d87-71a610c22b55"},
            {"nome_excel": "acompanhamento_comercial", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "1bc55ad3-6268-4ca8-9a65-43d645fde81f"},
            {"nome_excel": "acompanhamento_crm_360", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "5dd2ba6c-707e-4a9d-b6b4-7aeafb6ee19d"},
            {"nome_excel": "AdManager_Germanica", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "1bccefed-e7ce-45db-8089-124815454cd5"},
            {"nome_excel": "atualizacao_tab_precos", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "5d0a4ca1-1d64-495c-a74b-24006a7860ab"},
            {"nome_excel": "desempenho_comercial", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "a123005d-d9d7-4b34-9a03-83844cdb488f"},
            {"nome_excel": "mapa_calor_cidades", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "711c5e36-da52-4c21-bc22-3de26a7804ce"},
            {"nome_excel": "mapa_cnaes", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "7fabefc4-825f-4fb1-b3bc-255fa5eb6c9e"},
            {"nome_excel": "monitoramento_crowley", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "7fabefc4-825f-4fb1-b3bc-255fa5eb6c9e"},
            {"nome_excel": "monitoramento_mercado_publicitario_v2", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "2663ef67-d67c-4509-8fdd-301912817d4b"},
            {"nome_excel": "painel_executivos", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "813fe424-ba47-40c4-a4da-94bb8c4e5da3"},
            {"nome_excel": "Previsao_PBI", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "ebd6e739-21b0-4908-bcb7-31a6c083a7ff"},
            {"nome_excel": "Vendas Multimeios", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "233bf959-894f-4cc4-b943-65c98b487048"},
            {"nome_excel": "vendas_tv_global", "w_id": "0b73efc3-c3cc-45f1-9240-519e55b13de5", "d_id": "67dca079-c37e-4381-9286-26874ad79cc0"},
            {"nome_excel": "Calendário Corporativo PBI", "w_id": "43774005-3ad4-4a71-a12c-c70f24edf487", "d_id": "8f8a7ecf-a628-47cc-8f3a-d77bad0d2566"},
            {"nome_excel": "sumario_executivo_prd", "w_id": "6d38c256-3ede-43cc-83c5-3510a3e6bc0c", "d_id": "be8e915a-42b3-4597-a01b-1c38251b1abe"},
            {"nome_excel": "EP_Tickets", "w_id": "fad50932-b25d-41c1-b165-1c1d0af6b5ad", "d_id": "3c7a5b7e-dc81-42e8-bf73-da4e223934ca"},
            {"nome_excel": "Painel de Caixa", "w_id": "fad50932-b25d-41c1-b165-1c1d0af6b5ad", "d_id": "f783141f-ea38-445d-a50d-afb796859a04"},
            {"nome_excel": "Painel de Caixa Acionistas", "w_id": "fad50932-b25d-41c1-b165-1c1d0af6b5ad", "d_id": "f783141f-ea38-445d-a50d-afb796859a04"},
            {"nome_excel": "Painel de Vendas", "w_id": "fad50932-b25d-41c1-b165-1c1d0af6b5ad", "d_id": "dc1134d1-7739-48a6-a949-3d9d08a18000"},
            {"nome_excel": "Resultado_Eventos_v01", "w_id": "fad50932-b25d-41c1-b165-1c1d0af6b5ad", "d_id": "f06cc3d6-a882-449d-ab4a-fa1917bf893a"},
            {"nome_excel": "Sumario Executivo Verticais", "w_id": "fad50932-b25d-41c1-b165-1c1d0af6b5ad", "d_id": "9863ec0f-105e-4eff-8f0d-1ddba3e87070"},
            {"nome_excel": "CALENDARIOJORNALISMO", "w_id": "571a4a31-79dc-4124-91a1-4130706f9cc2", "d_id": "d6281474-49ba-4b09-a253-c13dbdf99158"},
            {"nome_excel": "painel_ga3_Esp_Publi", "w_id": "571a4a31-79dc-4124-91a1-4130706f9cc2", "d_id": "56cb2f15-6289-494e-a953-6c0c6a21c353"},
            {"nome_excel": "Painel Estratégico de Publicações Comerciais", "w_id": "13087102-ad42-4b34-8c98-67d29e2c4751", "d_id": "d655619b-fc37-407a-8568-efe6174326d9"},
            {"nome_excel": "mensuração_mkt", "w_id": "35450d52-832b-4557-98f5-f2f88da17c40", "d_id": "59d5e2cb-c0ab-4659-a618-d61e37f1542e"},
            {"nome_excel": "Desempenho - Entretenimento", "w_id": "1abc4333-9f61-43db-8926-59ed05976d2a", "d_id": "aafdde78-e019-45a2-8d0a-9227db57575b"},
            {"nome_excel": "Desempenho - Jornalismo Local", "w_id": "1abc4333-9f61-43db-8926-59ed05976d2a", "d_id": "aafdde78-e019-45a2-8d0a-9227db57575b"},
            {"nome_excel": "Desempenho Programação", "w_id": "1abc4333-9f61-43db-8926-59ed05976d2a", "d_id": "aafdde78-e019-45a2-8d0a-9227db57575b"},
            {"nome_excel": "Grade Diária", "w_id": "1abc4333-9f61-43db-8926-59ed05976d2a", "d_id": "aafdde78-e019-45a2-8d0a-9227db57575b"},
            {"nome_excel": "Grade_Diaria_v.2", "w_id": "1abc4333-9f61-43db-8926-59ed05976d2a", "d_id": "aafdde78-e019-45a2-8d0a-9227db57575b"},
            {"nome_excel": "Painel Gerencial - Promo e Ativação", "w_id": "1abc4333-9f61-43db-8926-59ed05976d2a", "d_id": "ec030553-9731-40b1-84af-06289ad86fc1"},
            {"nome_excel": "Orçamento_2026", "w_id": "4963782d-ca3c-4507-9fc1-6cfe1216c285", "d_id": "1fd2be51-36a9-4ede-ab5c-7d5820d5793c"},
            {"nome_excel": "painel_pessoas_prd_v1.1.0.pbix", "w_id": "4963782d-ca3c-4507-9fc1-6cfe1216c285", "d_id": "64341ff5-89d1-4012-a991-9546209ad04b"},
            {"nome_excel": "Remuneração_BI", "w_id": "4963782d-ca3c-4507-9fc1-6cfe1216c285", "d_id": "3d844e4e-4c66-4a45-9183-b67603882932"},
            {"nome_excel": "Unidade_Eventos", "w_id": "bd03572a-f34d-4ff6-9bb4-541d4ab5118c", "d_id": "ab31655b-e522-44f8-9bde-ad3a818858a0"},
            {"nome_excel": "powerbi_relatorios", "w_id": "c8250554-3e79-46cd-a860-7005e77301b1", "d_id": "e5c6d480-715e-4297-bc50-a801748a3680"},
            {"nome_excel": "Energia_GrupoEP_V4", "w_id": "838a4381-6a88-4235-a361-df7349246528", "d_id": "ddc95106-8ab9-4384-87d0-61517299ee40"},
            {"nome_excel": "Indicadores de Frotas_V4", "w_id": "838a4381-6a88-4235-a361-df7349246528", "d_id": "2fdf5b45-32a2-4b26-9d61-778620bae5c5"},
            {"nome_excel": "Indicadores de Frotas - ok", "w_id": "838a4381-6a88-4235-a361-df7349246528", "d_id": "eec9cbea-13fe-42f1-8a32-009f59ce9bd6"}
    ]

    print(f"🚀 Iniciando monitoramento de {len(meus_relatorios)} relatórios...")

    # 🔥 CARREGA EXCEL UMA ÚNICA VEZ
    estagio_robo.carregar_workbook()

    for item in meus_relatorios:
        status, detalhe, data_refresh, hora_refresh = estagio_robo.obter_status_e_detalhe(
            item["w_id"],
            item["d_id"],
            TOKEN
        )
        
        if data_refresh:
            estagio_robo.atualizar_catalogo_reports(
                item["nome_excel"],
                data_refresh,
                hora_refresh
            )

        estagio_robo.atualizar_monitoramento(item["nome_excel"], status)
        
        if status == "Erro":
            estagio_robo.registrar_erro_dashboard(item["nome_excel"], detalhe)

    # 🔥 SALVA UMA ÚNICA VEZ
    estagio_robo.salvar_workbook()

    print("🏁 Processo concluído com sucesso!")
