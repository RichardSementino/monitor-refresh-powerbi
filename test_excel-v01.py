import io
import openpyxl
from datetime import datetime
import locale
import requests  # Adicionado para futuras requisições à API
from azure.storage.blob import BlobServiceClient

# Configurando a data
try:
    locale.setlocale(locale.LC_TIME, "portuguese_brazil")
except:
    locale.setlocale(locale.LC_TIME, "pt_BR.utf8")

class MonitorDadosEPTV:
    def __init__(self, conn_str, container, arquivo):
        # O 'self' torna essas conexões acessíveis em qualquer parte da classe
        self.service_client = BlobServiceClient.from_connection_string(conn_str)
        self.blob_client = self.service_client.get_blob_client(container=container, blob=arquivo)

    def obter_status_powerbi(self, workspace_id, dataset_id, token_acesso):
        """
        Consulta a API do Power BI para saber se a última atualização foi OK.
        """
        print(f"📡 Consultando API do Power BI para o Dataset: {dataset_id}...")
        
        url = f"https://api.powerbi.com/v1.0/myorg/groups/{workspace_id}/datasets/{dataset_id}/refreshes?$top=1"
        
        headers = {
            "Authorization": f"Bearer {token_acesso}",
            "Content-Type": "application/json"
        }

        try:
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                dados = response.json()
                if dados['value']:
                    status_pbi = dados['value'][0]['status'] # 'Completed' ou 'Failed'
                    return "OK" if status_pbi == "Completed" else "Erro"
            return "Erro"
        except Exception as e:
            print(f"⚠️ Falha na comunicação com Power BI: {e}")
            return "Erro"

    def atualizar_monitoramento(self, painel, status):
        print(f"🔄 Iniciando atualização para o painel: {painel}...")
        
        # 1. Download do blob
        stream = self.blob_client.download_blob().readall()
        file_buffer = io.BytesIO(stream)
        
        # 2. Manipulação do Excel
        wb = openpyxl.load_workbook(file_buffer)
        ws = wb["Monitoramento Diário"]
        
        # 3. Localização Robusta
        # Pegamos apenas a DATA (ano, mês, dia) sem hora
        hoje_dt = datetime.now().date()
        
        # Geramos a string em português com BARRAS para o Log
        hoje_pt = hoje_dt.strftime("%d/%b/%y").lower()
        print(f"📅 Procurando a data de hoje: {hoje_pt}...")
        
        row_idx, col_idx = None, None

        # --- BUSCA DA LINHA (DATA) ---
        for r in range(5, ws.max_row + 1):
            cell = ws.cell(row=r, column=1)
            val = cell.value
            
            if val is None: continue

            # Tenta comparar como objeto de data (mais seguro)
            if isinstance(val, datetime):
                if val.date() == hoje_dt:
                    row_idx = r
                    break
            # Tenta comparar como string (caso o Excel trate como texto)
            else:
                # Remove espaços e padroniza para minúsculo
                txt_celula = str(val).strip().lower().replace("-", "/")
                if txt_celula == hoje_pt:
                    row_idx = r
                    break
        
        # --- BUSCA DA COLUNA (PAINEL) ---
        for c in range(2, ws.max_column + 1):
            header_val = ws.cell(row=4, column=c).value
            if header_val and str(header_val).strip() == painel:
                col_idx = c
                break

        # 4. Escrita e Upload
        if row_idx and col_idx:
            ws.cell(row=row_idx, column=col_idx).value = status
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            self.blob_client.upload_blob(output, overwrite=True)
            print(f"✅ Sucesso! Status '{status}' gravado na célula {ws.cell(row=row_idx, column=col_idx).coordinate}.")
        else:
            motivo = "Data não encontrada" if not row_idx else "Painel não encontrado"
            print(f"❌ Erro: {motivo} ({painel} em {hoje_pt}).")

# --- CONFIGURAÇÃO FINAL ---
if __name__ == "__main__":
    #key 1
    MINHA_CONEXAO = "DefaultEndpointsProtocol=https;AccountName=stmonitoramentoestudante;AccountKey=D+ItZxi5Bcl7p3mTdERCLj0hlzLykqHpFHD7aGGAo03MivmldhcbSgbb95FKmT81E1WhiKDv+62i+ASthGr67Q==;EndpointSuffix=core.windows.net"
    
    estagio_robo = MonitorDadosEPTV(MINHA_CONEXAO, "monitoramento", "Catálogos de Gestão_Dados_copia.xlsx")
    # 1. LISTA DE RELATÓRIOS (Adicione aqui os nomes exatos da linha 4 do seu Excel)
    paineis_para_atualizar = [
        "Painel Painéis Rodoviários",
        "Painel Compras",
        "Painel Contratos",
        "Performance Campanhas"
    ]
    
    # 2. LOOP AUTOMÁTICO
    print(f"🚀 Iniciando atualização de {len(paineis_para_atualizar)} painéis...")
    
    for painel in paineis_para_atualizar:
        
        estagio_robo.atualizar_monitoramento(painel, "OK")

    print("🏁 Todos os painéis foram processados!")
    